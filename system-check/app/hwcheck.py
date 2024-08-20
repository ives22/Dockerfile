#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Date   :2022/11/1
# @Author :p_pyjli
# @File   :hwcheck.py
# @Desc   :

import paramiko
import time
import os
from concurrent.futures import ThreadPoolExecutor
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill

# host_info = {
#     # "124.71.33.240": {"con_info": {"user": "root", "port": 22, "pass": "ives.123", "vm": "cvm", "type": "准入网关服务器", "cpu": 2, "mem": 8}, "cpu_info": {}, "mem_info": {},
#     #                   "disk_info": {}},
#     # "139.155.85.40": {"con_info": {"user": "root", "port": 22, "pass": "ives.123", "vm": "cvm", "type": "准入网关服务器", "cpu": 2, "mem": 8}, "cpu_info": {}, "mem_info": {},
#     #                   "disk_info": {}},
#     "120.24.222.91": {
#         "con_info": {"user": "root", "port": 22, "pass": "ives.123", "vm": "cvm", "type": "准入网关服务器", "cpu": 2,
#                      "mem": 8}, "cpu_info": {}, "mem_info": {},
#         "disk_info": {}},
# }
host_info = {
    "10.190.147.247": {"con_info": {"user": "root", "port": 22, "pass": "xxx", "本地vm": "cvm", "type": "接入网关服务器", "cpu": 4, "mem": 8}, "cpu_info": {}, "mem_info": {}, "disk_info": {}},
    "10.190.147.248": {"con_info": {"user": "root", "port": 22, "pass": "xxx", "本地vm": "cvm", "type": "接入网关服务器", "cpu": 4, "mem": 8}, "cpu_info": {}, "mem_info": {}, "disk_info": {}},
    "10.190.147.249": {"con_info": {"user": "root", "port": 22, "pass": "xxx", "本地vm": "cvm", "type": "接入网关服务器", "cpu": 4, "mem": 8}, "cpu_info": {}, "mem_info": {}, "disk_info": {}},
    "10.190.143.251": {"con_info": {"user": "root", "port": 22, "pass": "xxx", "本地vm": "cvm", "type": "准入/API网关服务器", "cpu": 8, "mem": 16}, "cpu_info": {}, "mem_info": {}, "disk_info": {}},
    "10.190.143.252": {"con_info": {"user": "root", "port": 22, "pass": "xxx", "本地vm": "cvm", "type": "准入/API网关服务器", "cpu": 8, "mem": 16}, "cpu_info": {}, "mem_info": {}, "disk_info": {}},
    "10.190.143.253": {"con_info": {"user": "root", "port": 22, "pass": "xxx", "本地vm": "cvm", "type": "准入/API网关服务器", "cpu": 8, "mem": 16}, "cpu_info": {}, "mem_info": {}, "disk_info": {}},
    "10.190.132.153": {"con_info": {"user": "root", "port": 22, "pass": "xxx", "本地vm": "cvm", "type": "日志分析服务器", "cpu": 4, "mem": 16}, "cpu_info": {}, "mem_info": {}, "disk_info": {}},
    "10.190.132.154": {"con_info": {"user": "root", "port": 22, "pass": "xxx", "本地vm": "cvm", "type": "日志分析服务器", "cpu": 4, "mem": 16}, "cpu_info": {}, "mem_info": {}, "disk_info": {}},
    "10.190.132.155": {"con_info": {"user": "root", "port": 22, "pass": "xxx", "本地vm": "cvm", "type": "日志分析服务器", "cpu": 4, "mem": 16}, "cpu_info": {}, "mem_info": {}, "disk_info": {}},
}

max_workers = 20
file_path = "./data"
file_name = "陕西人社项目-巡检报告"


def exec_cmd(host, user, password, port, cmd, pkeyt=False, timeout=20):
    ssh_client = paramiko.SSHClient()  # 创建一个客户端连接实例
    ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy)  # 加了这一句,如果第一次ssh连接要你输入yes,也不用输入了
    try:
        if pkeyt:
            # pkey = '/root/.ssh/id_rsa'
            pkey = '/Users/liyanjie/.ssh/id_rsa'
            key = paramiko.RSAKey.from_private_key_file(pkey)
            ssh_client.load_system_host_keys()
            ssh_client.connect(
                hostname=host,
                # port=port,
                username=user,
                pkey=key,
                # timeout=timeout
            )
        else:
            ssh_client.connect(
                hostname=host,
                port=port,
                username=user,
                password=password,
                timeout=timeout
            )
        stdin, stdout, stderr = ssh_client.exec_command(cmd)  # 执行一个命令,有标准输入,输出和错误输出
        stdour_ret = stdout.read()  # 标准输出赋值
        stderr_ret = stderr.read()  # 错误输出赋值
        result = stdour_ret if stdour_ret else stderr_ret  # 三元运算,或者换成判断语句.实现正确就标准输出赋值,错误就错误输出赋值
        result = result.decode('utf-8')
        ssh_client.close()
        ret = result.strip('\n')
        # ret = result
        return {"host": host, "ret": ret}
        # return ret  # 返回执行命令的服务器IP：命令结果
    except:
        print("连接失败")
        exec(1)


def call_back_disk(ret):
    for i in ret:
        host = i['host']
        data = i['ret']
        disk_info = {}
        for path in data.split('\n'):
            if path.split()[-1] == "/":
                disk_info["system_size"] = path.split()[1]  # 总大小
                disk_info["system_used"] = path.split()[-3]  # 使用多少G
                disk_info["system_use"] = path.split()[-2]  # 使用百分比
            else:
                disk_info["data_use"] = "/data: " + path.split()[-2]
        if host_info.get(host):
            host_info[host]["disk_info"] = disk_info


def call_back_cpu(ret):
    for i in ret:
        host = i['host']
        data = i['ret']
        cpu_info = {}
        cpu_info["load1"] = data.split('\n')[0].split()[1]
        cpu_info["load5"] = data.split('\n')[0].split()[2]
        cpu_info["load15"] = data.split('\n')[0].split()[3]
        cpu_info["id"] = data.split('\n')[1]
        host_info[host]["cpu_info"] = cpu_info
        host_info[host]["con_info"]["uptime"] = int(data.split('\n')[0].split()[0])


def call_back_mem(ret):
    for i in ret:
        host = i['host']
        data = i['ret']
        mem_info = {}
        mem_info['use'] = str(round(float(data), 2)) + "%"
        host_info[host]["mem_info"] = mem_info


def call_back_net(ret):
    cni = "eth0"
    for i in ret:
        host = i['host']
        data = i['ret']
        net_info = {}
        # net_info['received'] = cni + ":received " + str(round(float(data.split()[0]), 2)) + 'mbps'
        # net_info['sent'] = str(round(float(data.split()[1]), 2)) + 'mbps'
        net_info['net'] = cni + ":received " + str(
            round(float(data.split()[0]), 2)) + 'mbps' + '\n' + cni + ":sent " + str(
            round(float(data.split()[1]), 2)) + 'mbps'
        host_info[host]["net_info"] = net_info


def call_back_hostname(ret):
    for i in ret:
        host = i['host']
        host_info[host]["con_info"]["hostname"] = i["ret"]


def thread_exec(ip_info, cmd):
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_list = [executor.submit(exec_cmd, ip, v.get("con_info").get("user"), v.get("con_info").get("pass"),
                                       v.get("con_info").get("port"), cmd, pkeyt=True) for ip, v in ip_info.items()]
        return [x.result(timeout=120) for x in future_list]


def save_to_excel(path, data):
    WorkBook = Workbook()
    Sheet = WorkBook.active
    Sheet.title = 'Sheet1'
    TableTitle = ['网关', 'IP', '服务器类型', '主机名', '运行时长/天', 'CPU/核', '内存/GB', '根目录总量/G', '根目录使用量/G', '根目录使用率/%',
                  'CPU平均负载/1min',
                  'CPU平均负载/5min', 'CPU平均负载/15min', 'CPU空闲时间', f'内存使用率/%', '磁盘使用率/%', '网卡流量/Kbps']
    TitleColumn = {}  # 存放每个title值所对应的列{'IP': 'A', '主机名': 'B', '运行时长': 'C', 'CPU/核': 'D', '内存/GB': 'E', '根目录总量': 'F',...}
    AllHostItemValues = []  # 存放所有主机的监控项值 列表信息。

    # 维护表头，写入表头数据
    for row in range(len(TableTitle)):
        Col = row + 1
        Column = Sheet.cell(row=1, column=Col)  # 获取单元格的位置
        Column.value = TableTitle[row]  # 写入数据
        TitleCol = Column.coordinate.strip('1')  # 获取Title所在的列
        TitleColumn[TableTitle[row]] = TitleCol  # 加入到TitleColumn

    # data = {'124.71.33.240': {'con_info': {'user': 'root', 'port': 22, 'pass': 'ives.123', 'vm': 'cvm', 'type': '准入网关服务器', 'cpu': 2, 'mem': 8, 'uptime': 372, 'hostname': 'huawei'}, 'cpu_info': {'load1': '0.00', 'load5': '0.01', 'load15': '0.05', 'id': 'id,'}, 'mem_info': {'use': '27.49%'}, 'disk_info': {'system_size': '40G', 'system_used': '30G', 'system_use': '21%'}, 'net_info': {'net': 'eth0:received 0.07mbps\neth0:sent 0.07mbps'}}, '139.155.85.40': {'con_info': {'user': 'root', 'port': 22, 'pass': 'ives.123', 'vm': 'cvm', 'type': '准入网关服务器', 'cpu': 2, 'mem': 8, 'uptime': 362, 'hostname': 'tencent'}, 'cpu_info': {'load1': '0.22', 'load5': '0.18', 'load15': '0.13', 'id': '93.5'}, 'mem_info': {'use': '23.28%'}, 'disk_info': {'system_size': '79G', 'system_used': '49G', 'system_use': '36%', 'data_use': '/data36%'}, 'net_info': {'net': 'eth0:received 0.01mbps\neth0:sent 0.01mbps'}}, '120.24.222.91': {'con_info': {'user': 'root', 'port': 22, 'pass': 'ives.123', 'vm': 'cvm', 'type': '准入网关服务器', 'cpu': 2, 'mem': 8, 'uptime': 198, 'hostname': 'aly'}, 'cpu_info': {'load1': '0.00', 'load5': '0.00', 'load15': '0.00', 'id': '90.6'}, 'mem_info': {'use': '6.21%'}, 'disk_info': {'system_size': '79G', 'system_used': '49G', 'system_use': '36%', 'data_use': '/data36%'}, 'net_info': {'net': 'eth0:received 0.0mbps\neth0:sent 0.01mbps'}}}

    for k, v in data.items():
        HostItemValues = []  # 定义一个空列表，用于存放主机的监控项的值
        HostItemValues.append(v["con_info"]["vm"])
        HostItemValues.append(k)
        HostItemValues.append(v["con_info"]["type"])
        HostItemValues.append(v["con_info"]["hostname"])
        HostItemValues.append(v["con_info"]["uptime"])
        HostItemValues.append(v["con_info"]["cpu"])
        HostItemValues.append(v["con_info"]["mem"])
        HostItemValues.append(v["disk_info"]["system_size"])
        HostItemValues.append(v["disk_info"]["system_used"])
        HostItemValues.append(v["disk_info"]["system_use"])
        HostItemValues.append(v["cpu_info"]["load1"])
        HostItemValues.append(v["cpu_info"]["load5"])
        HostItemValues.append(v["cpu_info"]["load15"])
        HostItemValues.append(v["cpu_info"]["id"])
        HostItemValues.append(v["mem_info"]["use"])
        if v["disk_info"].get("data_use"):
            HostItemValues.append(v["disk_info"]["data_use"])
        else:
            HostItemValues.append('')
        HostItemValues.append(v["net_info"]["net"])

        # 将每一台主机的所有监控项信息添加到AllHostItems列表中
        AllHostItemValues.append(HostItemValues)

    #     print(HostItemValues)
    # print(AllHostItemValues)

    # 将所有信息写入到表格中
    pbar = tqdm(total=len(AllHostItemValues))
    for HostValue in range(len(AllHostItemValues)):
        pbar.update(1)
        Sheet.append(AllHostItemValues[HostValue])
    pbar.close()

    ############ 设置单元格样式 ############
    # 字体样式
    TitleFont = Font(name="宋体", size=12, bold=True, italic=False, color="000000")
    TableFont = Font(name="宋体", size=11, bold=False, italic=False, color="000000")
    # 对齐样式
    alignment = Alignment(horizontal="center", vertical="center", text_rotation=0, wrap_text=True)
    # 边框样式
    side1 = Side(style='thin', color='000000')
    border = Border(left=side1, right=side1, top=side1, bottom=side1)
    # 填充样式
    pattern_fill = PatternFill(fill_type='solid', fgColor='99ccff')
    # 设置列宽
    column_width = {'A': 12, 'B': 20, 'C': 30, 'D': 14, 'E': 10, 'F': 10, 'G': 10, 'H': 11, 'I': 11, 'J': 12, 'K': 14,
                    'L': 14, 'M': 14, 'N': 16, 'O': 16, 'P': 16, 'Q': 30}
    for i in column_width:
        Sheet.column_dimensions[i].width = column_width[i]
    # 设置首行的高度
    Sheet.row_dimensions[1].height = 50
    # 冻结窗口
    Sheet.freeze_panes = 'C2'
    # 添加筛选器
    Sheet.auto_filter.ref = Sheet.dimensions

    # 设置单元格字体及样式
    for row in Sheet.rows:
        for cell in row:
            if cell.coordinate.endswith('1') and len(cell.coordinate) == 2:
                cell.alignment = alignment  # 设置对齐样式
                cell.font = TitleFont  # 设置字体
                cell.border = border  # 设置边框样式
                cell.fill = pattern_fill  # 设置填充样式
            else:
                cell.font = TableFont
                cell.alignment = alignment
                cell.border = border

    WorkBook.save(filename=path)


def main():
    host_cmd = 'hostname'
    disk_cmd = 'df -h |grep  -E "/$|/data" |head -2'
    cpu_cmd = "uptime |awk '{print $3,$10,$11,$12}' | tr -d ',' && top -bn1 | grep Cpu | awk '{print $8}'"
    mem_cmd = "free -m  | grep Mem |awk '{print $3*100/$2}'"
    # net_cmd = """sar -n DEV 1 59 | grep Average |grep eth0| awk '{print$2,"\\t","input:","\\t",$5*1000*8/1024/1024,"mbps","\\n",$2,"\\t","output:","\\t",$6*1000*8/1024/1024,"mbps"}'"""
    net_cmd = """sar -n DEV 1 59 | grep Average |grep eth0| awk '{print $5*1000*8/1024/1024, $6*1000*8/1024/1024}'"""

    host_ret = thread_exec(ip_info=host_info, cmd=host_cmd)
    disk_ret = thread_exec(ip_info=host_info, cmd=disk_cmd)
    cpu_ret = thread_exec(ip_info=host_info, cmd=cpu_cmd)
    mem_ret = thread_exec(ip_info=host_info, cmd=mem_cmd)
    net_ret = thread_exec(ip_info=host_info, cmd=net_cmd)

    call_back_net(net_ret)
    call_back_mem(mem_ret)
    call_back_cpu(cpu_ret)
    call_back_disk(disk_ret)
    call_back_hostname(host_ret)

    date = time.strftime('%Y-%m-%d')
    fileName = os.path.join(file_path, file_name + date + '.xlsx')
    save_to_excel(fileName, host_info)


if __name__ == "__main__":
    while True:
        main()
        time.sleep(10800)


